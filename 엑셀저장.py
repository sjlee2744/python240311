import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import random

# 엑셀 파일 생성
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Products"

# 열 제목 추가
headers = ['제품ID', '제품 이름', '제품 가격']
for col_idx, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)

# 제품 정보 생성 및 저장
num_products = 100
for row_idx in range(2, num_products + 2):  # 2부터 시작하여 첫 번째 행은 제목이므로 건너뜁니다.
    product_id = row_idx - 1
    product_name = f"제품{product_id}"
    product_price = round(random.uniform(10, 1000), 2)  # 무작위 가격 생성

    # 엑셀에 데이터 추가
    sheet.cell(row=row_idx, column=1).value = product_id
    sheet.cell(row=row_idx, column=2).value = product_name
    sheet.cell(row=row_idx, column=3).value = product_price

# 파일 저장
file_name = "products.xlsx"
wb.save(file_name)
print(f"엑셀 파일 '{file_name}'이 생성되었습니다.")
